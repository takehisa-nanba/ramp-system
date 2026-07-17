import openpyxl
from io import BytesIO
from flask import Blueprint, request, send_file, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from backend.app.extensions import db
from backend.app.models import Supporter, SupporterTimecard, StaffDailyShift
from datetime import datetime
import calendar

export_bp = Blueprint('export', __name__, url_prefix='/api/management/export')

@export_bp.route('/attendance', methods=['GET'])
@jwt_required()
def export_attendance():
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    
    if not year or not month:
        now = datetime.now()
        year, month = now.year, now.month

    # 行政提出用のエクセルを新規作成（本来はテンプレートをロードする）
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月勤務実績表"
    
    # ヘッダー
    ws.cell(row=1, column=1, value="職員名")
    ws.cell(row=1, column=2, value="職種")
    
    _, last_day = calendar.monthrange(year, month)
    
    for day in range(1, last_day + 1):
        ws.cell(row=1, column=2 + day, value=f"{month}/{day}")
        
    supporters = Supporter.query.filter_by(is_active=True).all()
    
    for row_idx, supporter in enumerate(supporters, start=2):
        ws.cell(row=row_idx, column=1, value=f"{supporter.last_name} {supporter.first_name}")
        roles = [r.name for r in supporter.roles]
        ws.cell(row=row_idx, column=2, value=",".join(roles) if roles else "一般")
        
        # 月の全データを取得
        timecards = SupporterTimecard.query.filter(
            SupporterTimecard.supporter_id == supporter.id,
            db.func.extract('year', SupporterTimecard.work_date) == year,
            db.func.extract('month', SupporterTimecard.work_date) == month
        ).all()
        
        for day in range(1, last_day + 1):
            target_date = datetime(year, month, day).date()
            tc = next((t for t in timecards if t.work_date == target_date), None)
            
            val = ""
            if tc and tc.check_in and tc.check_out:
                in_str = tc.check_in.strftime('%H:%M')
                out_str = tc.check_out.strftime('%H:%M')
                val = f"{in_str}-{out_str}"
            elif tc and tc.is_absent:
                val = "欠勤"
                
            ws.cell(row=row_idx, column=2 + day, value=val)
            
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"attendance_report_{year}_{month:02d}.xlsx"
    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
